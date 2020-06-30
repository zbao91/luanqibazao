# encoding: utf-8
"""
This Script is for excel processing
"""
import xlrd
import openpyxl
import numpy as np

from .base import List

class ExlProcess():
    def __init__(self):
        self.list_obj = List()
        pass

    def extract_data(self, excel_file, sheet_name=None, way='all'):
        """
        get data from sheet by index or name of sheet.
        - excel_file:
            * describe: excel file path
            * type: str
        - sheet:
            * describe: which sheet be loaded, index of sheet/ name of sheet
            * type: str(for name) or int(for index ** start from 0 **)
        - way:
            * describe: how to extract data, by index or name or all
            * type: str
        ref: https://stackoverflow.com/questions/33522969/openpyxl-fetch-value-from-excel-and-store-in-key-value-pair
        """
        # in case, only one sheet is enter
        if isinstance(sheet_name, str) or isinstance(sheet_name, int):
            sheet_name = [sheet_name]
        # load files
        book = openpyxl.load_workbook(excel_file)
        # sheet to iterate
        all_sheets = book.sheetnames
        if way == 'name':
            sheets_name = self.list_obj.get_common(all_sheets, sheet_name)
        elif way == 'index':
            sheets_name = self.list_obj.get_by_multi_indices(all_sheets, sheet_name)
            # when there is only one sheet_name entered, the returned value will be string
            if isinstance(sheets_name, str):
                sheets_name = [sheets_name]
        else:
            sheets_name = all_sheets

        data = {}
        for s_name in sheets_name:
            data[s_name] = []
            activate = book.get_sheet_by_name(s_name)
            # iterate sheet
            res = self.iter_rows(activate)
            keys = next(res)
            for new in res:
                row = dict(zip(keys, new))
                data[s_name].append(row)
        return data

    def iter_rows(self, sheet):
        """
            iterate rows in sheet
        """
        for row in sheet.iter_rows():
            yield [cell.value for cell in row]

    def write_data(self, excel_file, sheet_name, data):
        """
            not finish yet !!!!! TODO: 还没完成
            write data to sheet
            ref: https://codereview.stackexchange.com/questions/227698/python-list-of-dictionaries-to-xls
        """
        book = openpyxl.load_workbook(excel_file)
        sheetnames = book.sheetnames
        sheet_name = sheetnames[sheet_name]
        sheet = book.get_sheet_by_name(sheet_name)
        headers = list(data[0])
        sheet.append(headers)
        for x in data:
            sheet.append(list(x.values()))
        book.save(excel_file)
        return